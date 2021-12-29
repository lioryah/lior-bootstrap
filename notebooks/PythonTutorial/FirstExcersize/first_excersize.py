import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']  # same as sheet.cell(1,1)



print(cell.value)


num_of_rows = sheet.max_row
print(num_of_rows)


for row in range(2, num_of_rows + 1):  # start from the second row
    print(row)    
    cell = sheet.cell(row, 3)
    print(cell.value)
    corrected_price  = cell.value * 0.9
    print(cell.value)

    '''
    if row == 1:
        continue
    sheet['c' + str(row)].value = sheet['c' + str(row)].value * 0.9
'''  